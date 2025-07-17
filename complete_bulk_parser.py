import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import time
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class CompleteBulkHochmaParser:
    def __init__(self):
        self.base_url = "https://nocr.net/com_kor_hochma"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.results = []
        self.failed_ids = []
        self.stats = {
            'total_articles': 0,
            'successful_parses': 0,
            'failed_parses': 0,
            'total_verses': 0,
            'start_time': None,
            'end_time': None
        }
    
    def detect_verse_pattern(self, content):
        """게시글의 절 구분 패턴 감지"""
        
        # 패턴 1: ====31:1 또는 ===31:1 형태
        equals_pattern = re.findall(r'={3,}(\d+):(\d+)', content)
        if equals_pattern:
            return 'equals', equals_pattern
        
        # 패턴 2: 줄바꿈 기반 31:1 형태
        lines = content.split('\n')
        line_pattern = []
        for i, line in enumerate(lines):
            line = line.strip()
            # 절 번호 패턴: 숫자:숫자 형태가 독립적인 줄에 있는 경우
            if re.match(r'^\d+:\d+([,-]\d+)*$', line):
                line_pattern.append((line, i))
        
        if line_pattern:
            return 'lines', line_pattern
        
        return 'none', []
    
    def parse_verse_reference(self, verse_ref):
        """절 참조를 파싱해서 개별 절 번호들로 변환"""
        verses = []
        
        # 콤마로 분리된 경우: 19:37,38
        if ',' in verse_ref:
            parts = verse_ref.split(',')
            chapter = parts[0].split(':')[0]
            for part in parts:
                if ':' in part:
                    verses.append(part.strip())
                else:
                    verses.append(f"{chapter}:{part.strip()}")
        
        # 하이픈으로 범위 지정된 경우: 19:33-35
        elif '-' in verse_ref and ':' in verse_ref:
            chapter_verse, end_verse = verse_ref.split('-')
            chapter, start_verse = chapter_verse.split(':')
            start_verse = int(start_verse)
            end_verse = int(end_verse.strip())
            
            for v in range(start_verse, end_verse + 1):
                verses.append(f"{chapter}:{v}")
        
        # 단일 절: 19:1
        else:
            verses.append(verse_ref.strip())
        
        return verses
    
    def parse_single_article(self, article_id):
        """단일 게시글 파싱"""
        url = f"{self.base_url}/{article_id}"
        
        try:
            response = requests.get(url, headers=self.headers, timeout=10)
            response.raise_for_status()
            
            # BeautifulSoup으로 파싱
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 제목 추출
            title = ""
            title_tag = soup.find('h1')
            if title_tag:
                title = title_tag.get_text(strip=True)
            else:
                title_tag = soup.find('title')
                if title_tag:
                    title = title_tag.get_text(strip=True)
            
            # 제목에서 정보 추출
            commentary_name = ""
            book_name = ""
            chapter = ""
            
            if title:
                # 주석 이름 추출
                if "주석" in title:
                    commentary_match = re.search(r'(\w+)\s*주석', title)
                    if commentary_match:
                        commentary_name = commentary_match.group(1) + " 주석"
                
                # 성경책 이름과 장 추출
                bible_books = ['창세기', '출애굽기', '레위기', '민수기', '신명기', '여호수아', '사사기', 
                              '룻기', '사무엘상', '사무엘하', '열왕기상', '열왕기하', '역대상', '역대하', 
                              '에스라', '느헤미야', '에스더', '욥기', '시편', '잠언', '전도서', '아가', 
                              '이사야', '예레미야', '예레미야애가', '에스겔', '다니엘', '호세아', '요엘', 
                              '아모스', '오바댜', '요나', '미가', '나훔', '하박국', '스바냐', '학개', 
                              '스가랴', '말라기', '마태복음', '마가복음', '누가복음', '요한복음', 
                              '사도행전', '로마서', '고린도전서', '고린도후서', '갈라디아서', '에베소서', 
                              '빌립보서', '골로새서', '데살로니가전서', '데살로니가후서', '디모데전서', 
                              '디모데후서', '디도서', '빌레몬서', '히브리서', '야고보서', '베드로전서', 
                              '베드로후서', '요한1서', '요한2서', '요한3서', '유다서', '요한계시록',
                              '요한일서', '요한이서', '요한삼서']
                
                for book in bible_books:
                    if book in title:
                        book_name = book
                        # 장 번호 추출
                        chapter_match = re.search(f'{book}\\s*(\\d+)장', title)
                        if chapter_match:
                            chapter = chapter_match.group(1)
                        break
            
            # 본문 내용 추출 (BR 태그를 줄바꿈으로 변환)
            content = ""
            content_selectors = ['.xe_content', '.rd_body', '.rhymix_content', '.document_content']
            
            for selector in content_selectors:
                content_div = soup.select_one(selector)
                if content_div:
                    # <br> 태그를 \n으로 변환
                    for br in content_div.find_all('br'):
                        br.replace_with('\n')
                    content = content_div.get_text()
                    break
            
            if not content:
                print(f"⚠️  {article_id}: 본문 내용을 찾을 수 없음")
                return []
            
            # 절 패턴 감지 및 파싱
            pattern_type, pattern_data = self.detect_verse_pattern(content)
            
            parsed_verses = []
            
            if pattern_type == 'equals':
                # === 또는 ==== 패턴
                content_parts = re.split(r'={3,}\d+:\d+', content)
                
                for i, (chapter_num, verse_num) in enumerate(pattern_data):
                    if i + 1 < len(content_parts):
                        verse_content = content_parts[i + 1].strip()
                        if verse_content:
                            parsed_verses.append({
                                'article_id': article_id,
                                'url': url,
                                'title': title,
                                'commentary_name': commentary_name,
                                'book_name': book_name,
                                'chapter': int(chapter_num),
                                'verse': int(verse_num),
                                'content': verse_content,
                                'content_length': len(verse_content),
                                'pattern_type': 'equals'
                            })
            
            elif pattern_type == 'lines':
                # 줄바꿈 기반 패턴
                lines = content.split('\n')
                
                for i, (verse_ref, line_idx) in enumerate(pattern_data):
                    # 다음 절 참조까지의 내용 추출
                    start_idx = line_idx + 1
                    end_idx = len(lines)
                    
                    if i + 1 < len(pattern_data):
                        end_idx = pattern_data[i + 1][1]
                    
                    verse_content = '\n'.join(lines[start_idx:end_idx]).strip()
                    
                    if verse_content:
                        # 절 참조 파싱 (콤마, 하이픈 처리)
                        individual_verses = self.parse_verse_reference(verse_ref)
                        
                        for verse_str in individual_verses:
                            if ':' in verse_str:
                                chapter_num, verse_num = verse_str.split(':')
                                parsed_verses.append({
                                    'article_id': article_id,
                                    'url': url,
                                    'title': title,
                                    'commentary_name': commentary_name,
                                    'book_name': book_name,
                                    'chapter': int(chapter_num),
                                    'verse': int(verse_num),
                                    'content': verse_content,
                                    'content_length': len(verse_content),
                                    'pattern_type': 'lines'
                                })
            
            else:
                # 패턴이 없는 경우 전체를 하나의 항목으로
                parsed_verses.append({
                    'article_id': article_id,
                    'url': url,
                    'title': title,
                    'commentary_name': commentary_name,
                    'book_name': book_name,
                    'chapter': int(chapter) if chapter else 0,
                    'verse': 0,
                    'content': content.strip(),
                    'content_length': len(content.strip()),
                    'pattern_type': 'none'
                })
            
            print(f"✅ {article_id}: {len(parsed_verses)}개 절 파싱 완료 ({pattern_type} 패턴)")
            return parsed_verses
            
        except Exception as e:
            print(f"❌ {article_id}: 파싱 실패 - {str(e)}")
            self.failed_ids.append(article_id)
            return []
    
    def parse_all_articles(self, article_ids):
        """모든 게시글 파싱"""
        print(f"🚀 {len(article_ids)}개 게시글 대량 파싱 시작...")
        
        self.stats['total_articles'] = len(article_ids)
        self.stats['start_time'] = datetime.now()
        
        for i, article_id in enumerate(article_ids, 1):
            print(f"\n[{i}/{len(article_ids)}] 파싱 중: {article_id}")
            
            verses = self.parse_single_article(article_id)
            
            if verses:
                self.results.extend(verses)
                self.stats['successful_parses'] += 1
                self.stats['total_verses'] += len(verses)
            else:
                self.stats['failed_parses'] += 1
            
            # 진행 상황 출력
            if i % 50 == 0:
                success_rate = (self.stats['successful_parses'] / i) * 100
                elapsed = datetime.now() - self.stats['start_time']
                print(f"📊 진행률: {i}/{len(article_ids)} ({i/len(article_ids)*100:.1f}%) | "
                      f"성공률: {success_rate:.1f}% | 절 수: {self.stats['total_verses']} | "
                      f"경과시간: {elapsed}")
            
            # 서버 부하 방지
            time.sleep(0.1)
        
        self.stats['end_time'] = datetime.now()
        
        print(f"\n🎉 파싱 완료!")
        print(f"  성공: {self.stats['successful_parses']}/{self.stats['total_articles']}")
        print(f"  실패: {self.stats['failed_parses']}/{self.stats['total_articles']}")
        print(f"  총 절 수: {self.stats['total_verses']}")
        
        duration = self.stats['end_time'] - self.stats['start_time']
        print(f"  소요 시간: {duration}")
    
    def save_to_excel(self, filename=None):
        """결과를 엑셀 파일로 저장"""
        if not self.results:
            print("저장할 데이터가 없습니다.")
            return
        
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"hochma_complete_commentaries_{timestamp}.xlsx"
        
        print(f"💾 엑셀 파일 저장 중: {filename}")
        
        # 데이터프레임 생성
        df = pd.DataFrame(self.results)
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "호크마 주석 데이터"
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터프레임을 엑셀에 쓰기
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 열 너비 조정
        column_widths = {
            'A': 12,  # article_id
            'B': 50,  # url
            'C': 40,  # title
            'D': 15,  # commentary_name
            'E': 15,  # book_name
            'F': 8,   # chapter
            'G': 8,   # verse
            'H': 80,  # content
            'I': 12,  # content_length
            'J': 12   # pattern_type
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 요약 시트 추가
        summary_ws = wb.create_sheet(title="요약 정보")
        
        summary_data = [
            ["항목", "값"],
            ["파싱 일시", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["총 게시글 수", self.stats['total_articles']],
            ["성공한 게시글", self.stats['successful_parses']],
            ["실패한 게시글", self.stats['failed_parses']],
            ["총 절 수", self.stats['total_verses']],
            ["파싱 시간", str(self.stats['end_time'] - self.stats['start_time']) if self.stats['end_time'] else ""],
            ["", ""],
            ["성경책별 통계", ""],
        ]
        
        # 성경책별 통계
        if len(df) > 0:
            book_stats = df.groupby('book_name').agg({
                'article_id': 'nunique',
                'verse': 'count',
                'content_length': 'sum'
            }).reset_index()
            
            for _, row in book_stats.iterrows():
                summary_data.append([
                    f"  {row['book_name']}", 
                    f"게시글 {row['article_id']}개, 절 {row['verse']}개, 총 {row['content_length']:,}자"
                ])
        
        for row in summary_data:
            summary_ws.append(row)
        
        # 요약 시트 스타일링
        for cell in summary_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 50
        
        # 파일 저장
        wb.save(filename)
        
        print(f"✅ 저장 완료: {filename}")
        print(f"   - 주석 데이터: {len(self.results)}행")
        if len(df) > 0:
            print(f"   - 요약 정보: 성경책 {len(book_stats)}개")
        
        return filename

def main():
    """메인 함수"""
    parser = CompleteBulkHochmaParser()
    
    # 1. 최신 JSON 파일 사용
    print("1️⃣ 게시글 ID 수집...")
    
    # 완전한 ID 목록 파일 찾기
    json_files = [f for f in os.listdir('.') if f.startswith('hochma_complete_ids_') and f.endswith('.json')]
    
    if json_files:
        # 가장 최근 파일 사용
        latest_file = sorted(json_files)[-1]
        print(f"   최신 완전 ID 파일 사용: {latest_file}")
        
        with open(latest_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            article_ids = data['article_ids']
    else:
        print("   완전한 ID 파일을 찾을 수 없습니다.")
        return
    
    print(f"   수집된 게시글: {len(article_ids)}개")
    
    # 2. 모든 게시글 파싱
    print("\n2️⃣ 모든 게시글 파싱...")
    parser.parse_all_articles(article_ids)
    
    # 3. 엑셀 파일로 저장
    print("\n3️⃣ 엑셀 파일 저장...")
    filename = parser.save_to_excel()
    
    print(f"\n🎉 모든 작업 완료!")
    print(f"📁 결과 파일: {filename}")
    
    # 실패한 ID들 출력
    if parser.failed_ids:
        print(f"\n⚠️  실패한 게시글 ID: {parser.failed_ids}")

if __name__ == "__main__":
    main() 